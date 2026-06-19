"""Экспорт результатов эксперимента в Excel (results.xlsx).

Читает results.json и формирует наглядную таблицу для ручного анализа:
- лист «Детально» — одна строка на задачу, по каждому методу видны ответ, верность,
  время и токены (для RLM ещё итерации и причина остановки);
- лист «Сводка» — агрегаты accuracy/время/токены по (метод × тип).

    python -m eval.report                 # из eval.results_path конфига
    python -m eval.report results.json out.xlsx
"""

from __future__ import annotations

import json
import sys
from collections import defaultdict
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from rlm.config import load_config

METHOD_ORDER = ["rlm", "rag", "naive"]
METHOD_TITLE = {"rlm": "RLM", "rag": "RAG", "naive": "Naive"}

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_METHOD_FILLS = {  # лёгкая заливка групп колонок по методам
    "rlm": PatternFill("solid", fgColor="E2EFDA"),
    "rag": PatternFill("solid", fgColor="FCE4D6"),
    "naive": PatternFill("solid", fgColor="DDEBF7"),
}
_OK_FONT = Font(color="2E7D32", bold=True)
_BAD_FONT = Font(color="C62828", bold=True)


def _methods_present(records: list[dict]) -> list[str]:
    present = {r["method"] for r in records}
    return [m for m in METHOD_ORDER if m in present] + sorted(present - set(METHOD_ORDER))


def _build_detail_sheet(wb: Workbook, records: list[dict], methods: list[str]) -> None:
    ws = wb.create_sheet("Детально")

    # Группируем записи по задаче.
    by_task: dict[str, dict[str, Any]] = defaultdict(dict)
    meta: dict[str, dict[str, Any]] = {}
    for r in records:
        by_task[r["task_id"]][r["method"]] = r
        meta.setdefault(r["task_id"], {
            "type": r["type"], "char_len": r["char_len"],
            "question": r.get("question", ""), "gold": r["gold"],
        })

    # Шапка: общие колонки + по 4–6 колонок на метод.
    base_cols = ["Задача", "Тип", "Длина, симв.", "ВОПРОС", "ПРАВИЛЬНЫЙ ОТВЕТ"]
    header: list[str] = list(base_cols)
    for m in methods:
        t = METHOD_TITLE.get(m, m)
        header += [f"{t}: ответ", f"{t}: верно", f"{t}: время,с", f"{t}: токены"]
        if m == "rlm":
            header += ["RLM: итераций", "RLM: остановка"]
    ws.append(header)

    for task_id in sorted(by_task, key=lambda k: (meta[k]["type"], meta[k]["char_len"])):
        mt = meta[task_id]
        row: list[Any] = [task_id, mt["type"], mt["char_len"], mt["question"], mt["gold"]]
        for m in methods:
            rec = by_task[task_id].get(m, {})
            row.append(rec.get("answer", "—"))
            row.append("✓" if rec.get("correct") else "✗")
            row.append(rec.get("elapsed", ""))
            row.append(rec.get("usage", {}).get("total_tokens", ""))
            if m == "rlm":
                row.append(rec.get("iterations", ""))
                row.append(rec.get("stopped_reason", ""))
        ws.append(row)

    _style_table(ws, header, methods)


def _style_table(ws, header: list[str], methods: list[str]) -> None:
    # Стиль шапки.
    for c, title in enumerate(header, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ширины колонок и перенос текста для «ответ»/«вопрос».
    widths = {"Задача": 12, "Тип": 13, "Длина, симв.": 12, "ВОПРОС": 48, "ПРАВИЛЬНЫЙ ОТВЕТ": 18}
    for c, title in enumerate(header, 1):
        letter = get_column_letter(c)
        if title in widths:
            ws.column_dimensions[letter].width = widths[title]
        elif title.endswith("ответ"):
            ws.column_dimensions[letter].width = 46
        elif "остановка" in title:
            ws.column_dimensions[letter].width = 15
        else:
            ws.column_dimensions[letter].width = 12

    # Заливка групп по методам + выравнивание/цвет верности по строкам.
    for row in ws.iter_rows(min_row=2):
        for c, cell in enumerate(row, 1):
            title = header[c - 1]
            for m in methods:
                if title.startswith(METHOD_TITLE.get(m, m) + ":"):
                    cell.fill = _METHOD_FILLS.get(m, PatternFill())
            if title.endswith("ответ") or title == "ВОПРОС":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if title.endswith("верно"):
                cell.alignment = Alignment(horizontal="center")
                cell.font = _OK_FONT if cell.value == "✓" else _BAD_FONT

    ws.freeze_panes = "F2"  # фиксируем общие колонки и шапку
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"


def _build_summary_sheet(wb: Workbook, records: list[dict], methods: list[str]) -> None:
    ws = wb.create_sheet("Сводка", 0)  # первой
    agg: dict[tuple, dict[str, float]] = defaultdict(lambda: {"n": 0, "ok": 0, "secs": 0.0, "tok": 0})
    types = []
    for r in records:
        key = (r["method"], r["type"])
        a = agg[key]
        a["n"] += 1
        a["ok"] += 1 if r["correct"] else 0
        a["secs"] += r.get("elapsed", 0) or 0
        a["tok"] += r.get("usage", {}).get("total_tokens", 0) or 0
        if r["type"] not in types:
            types.append(r["type"])

    header = ["Метод", "Тип", "Точность", "Верно/Всего", "Ср. время, с", "Ср. токены"]
    ws.append(header)
    for m in methods:
        for t in types:
            a = agg.get((m, t))
            if not a:
                continue
            ws.append([
                METHOD_TITLE.get(m, m), t, round(a["ok"] / a["n"], 3),
                f"{int(a['ok'])}/{int(a['n'])}",
                round(a["secs"] / a["n"], 1), int(a["tok"] / a["n"]),
            ])
    # Итог по методам.
    ws.append([])
    by_m: dict[str, list[int]] = defaultdict(lambda: [0, 0])
    for r in records:
        by_m[r["method"]][0] += 1
        by_m[r["method"]][1] += 1 if r["correct"] else 0
    for m in methods:
        n, ok = by_m[m][0], by_m[m][1]
        ws.append([f"ИТОГО {METHOD_TITLE.get(m, m)}", "", round(ok / n, 3) if n else 0, f"{ok}/{n}", "", ""])

    for c, _ in enumerate(header, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for c, w in enumerate([16, 14, 10, 14, 14, 12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    # Точность в процентном формате.
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0%"


_LEADER_FILL = PatternFill("solid", fgColor="C6EFCE")
_LEADER_FONT = Font(bold=True, color="006100")


def _build_by_length_sheet(wb: Workbook, records: list[dict], methods: list[str]) -> None:
    """Срез точности по длине контекста — главная иллюстрация независимости RLM от длины."""
    ws = wb.create_sheet("По длинам", 1)  # после «Сводки»

    # Бакеты длин из конфига; char_len относим к ближайшей цели.
    buckets = sorted(load_config()["eval"]["context_lengths"])

    def bucket_of(char_len: int) -> int:
        return min(buckets, key=lambda b: abs(b - char_len))

    # Порядок типов — как встречаются в данных.
    types: list[str] = []
    for r in records:
        if r["type"] not in types:
            types.append(r["type"])

    # agg[(тип|"*", бакет, метод)] = [n, ok]
    agg: dict[tuple, list[int]] = defaultdict(lambda: [0, 0])
    for r in records:
        b = bucket_of(r["char_len"])
        ok = 1 if r["correct"] else 0
        for t in (r["type"], "*"):  # "*" = агрегат «Все типы»
            agg[(t, b, r["method"])][0] += 1
            agg[(t, b, r["method"])][1] += ok

    header = ["Тип задачи", "Длина"] + [METHOD_TITLE.get(m, m) for m in methods]
    ws.append(header)

    def emit_section(type_key: str, title: str) -> None:
        for b in buckets:
            present = [agg.get((type_key, b, m)) for m in methods]
            if not any(present):
                continue
            accs = []
            row: list[Any] = [title, f"{round(b / 1000)}k"]
            for m in methods:
                cell = agg.get((type_key, b, m))
                if cell and cell[0]:
                    acc = cell[1] / cell[0]
                    accs.append(acc)
                    row.append(f"{acc:.0%} ({cell[1]}/{cell[0]})")
                else:
                    accs.append(None)
                    row.append("—")
            ws.append(row)
            # Подсветка лидера(ов) среди методов в этой строке.
            best = max([a for a in accs if a is not None], default=None)
            if best is not None:
                r_idx = ws.max_row
                for j, a in enumerate(accs):
                    if a is not None and abs(a - best) < 1e-9:
                        c = ws.cell(row=r_idx, column=3 + j)
                        c.fill = _LEADER_FILL
                        c.font = _LEADER_FONT
            title = ""  # имя типа печатаем только в первой строке секции

    for t in types:
        emit_section(t, t)
    ws.append([])
    emit_section("*", "ВСЕ ТИПЫ")

    # Стиль шапки и ширины.
    for c, _ in enumerate(header, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 8
    for c in range(3, 3 + len(methods)):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "C2"


def write_report(results_path: str, xlsx_path: str) -> str:
    with open(results_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    records = data["records"]
    methods = _methods_present(records)

    wb = Workbook()
    wb.remove(wb.active)  # убрать дефолтный пустой лист
    _build_summary_sheet(wb, records, methods)
    _build_by_length_sheet(wb, records, methods)
    _build_detail_sheet(wb, records, methods)
    wb.save(xlsx_path)
    return xlsx_path


def main() -> None:
    cfg = load_config()
    results_path = sys.argv[1] if len(sys.argv) > 1 else cfg["eval"]["results_path"]
    xlsx_path = sys.argv[2] if len(sys.argv) > 2 else results_path.rsplit(".", 1)[0] + ".xlsx"
    path = write_report(results_path, xlsx_path)
    print(f"Excel-отчёт сохранён: {path}")


if __name__ == "__main__":
    main()
